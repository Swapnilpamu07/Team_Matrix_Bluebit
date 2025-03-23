import pyfiglet  # type: ignore
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import time
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager  # type: ignore

# ✅ Display Stylish Header
ascii_art = pyfiglet.figlet_format("WEBZ-SCRAPPER\nCREATED BY\nTeamMatrix", font="slant")
print("\033[1;36m" + ascii_art + "\033[0m")  # Cyan color text

# ✅ Get User Input
job_role = input("Enter Job Role (e.g., Software Engineer): ").strip().replace(" ", "%20")
job_location = input("Enter Job Location (e.g., India): ").strip().replace(" ", "%20")
job_field = input("Enter Job Field (e.g., IT, Finance, Healthcare): ").strip().replace(" ", "%20")

# ✅ Generate Filename with Date
current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
SAVE_PATH = f"C:/Users/R5 5500U/Aditya/Video/OneShot/Team_Matrix_Bluebit/ai-ml/data_path/{job_role}{job_location}{current_datetime}.xlsx"

# ✅ Job Search URLs
LINKEDIN_URL = f"https://www.linkedin.com/jobs/search/?keywords={job_role}%20{job_field}&location={job_location}"
NAUKRI_URL = f"https://www.naukri.com/{job_role.replace('%20','-')}-jobs-in-{job_location.replace('%20','-')}"

# ✅ Setup Selenium WebDriver
options = Options()
options.add_argument("--start-maximized")
options.add_argument("--headless")  # Run without UI
options.add_argument("--disable-gpu")  # Fix GPU error
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.implicitly_wait(10)  # ✅ Wait for elements to load

# ✅ Scrape LinkedIn Jobs
print("🔎 Scraping LinkedIn Jobs...")
driver.get(LINKEDIN_URL)
time.sleep(5)

linked_jobs = []
jobs = driver.find_elements(By.CSS_SELECTOR, ".jobs-search-results__list-item")  # ✅ Updated selector

for job in jobs[:10]:  # Scraping first 10 jobs
    try:
        title = job.find_element(By.CSS_SELECTOR, ".base-search-card__title").text.strip()
        company = job.find_element(By.CSS_SELECTOR, ".base-search-card__subtitle").text.strip()
        location = job.find_element(By.CSS_SELECTOR, ".job-search-card__location").text.strip()
        job_link = job.find_element(By.TAG_NAME, "a").get_attribute("href")
        linked_jobs.append(["LinkedIn", title, company, location, job_link])
    except Exception as e:
        print(f"⚠ LinkedIn Error: {e}")
        continue

# ✅ Scrape Naukri Jobs
print("🔎 Scraping Naukri Jobs...")
driver.get(NAUKRI_URL)
time.sleep(5)

naukri_jobs = []
jobs = driver.find_elements(By.CLASS_NAME, "jobTuple")  

for job in jobs[:10]:  # Scraping first 10 jobs
    try:
        title = job.find_element(By.CLASS_NAME, "title").text.strip()
        company = job.find_element(By.CLASS_NAME, "companyName").text.strip()
        location = job.find_element(By.CLASS_NAME, "location").text.strip()
        job_link = job.find_element(By.TAG_NAME, "a").get_attribute("href")
        naukri_jobs.append(["Naukri", title, company, location, job_link])
    except Exception as e:
        print(f"⚠ Naukri Error: {e}")
        continue

driver.quit()

# ✅ Combine Data
all_jobs = linked_jobs + naukri_jobs
df = pd.DataFrame(all_jobs, columns=["Source", "Title", "Company", "Location", "Job Link"])

# ✅ Ensure Save Path Exists
os.makedirs(os.path.dirname(SAVE_PATH), exist_ok=True)

# ✅ Save to Excel with Stylish Formatting
with pd.ExcelWriter(SAVE_PATH, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name="Job Listings", index=False)
    
    # ✅ Load Workbook and Worksheet
    workbook = writer.book
    worksheet = writer.sheets["Job Listings"]
    
    # ✅ Styling Headers
    header_font = Font(bold=True, color="FFFFFF")  # White text
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue header
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    # ✅ Apply Alternating Row Colors
    light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(2, len(df) + 2):  
        if row % 2 == 0:  
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=row, column=col).fill = light_gray
    
    # ✅ Auto-Adjust Column Width
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        worksheet.column_dimensions[col_letter].width = max_length + 3

print(f"\n✅ Scraping complete! Data saved to: {SAVE_PATH}")
print("🔥 Open the file to check your results! 🔥")